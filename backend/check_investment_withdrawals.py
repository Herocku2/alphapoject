#!/usr/bin/env python
"""
Script para verificar inversiones vs retiros de cada usuario y generar Excel
Ejecutar con: python manage.py shell < check_investment_withdrawals.py
O dentro de Django shell: exec(open('check_investment_withdrawals.py').read())
"""

from apps.investment_plans.models import UserInvestment, InvestmentPlanTransaction
from apps.investment_plans.choices import TransactionStatus
from apps.withdrawals.models import Withdrawal
from apps.withdrawals.choices import WithdrawalType
from apps.transfers.models import UserTransfer
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

print("=" * 80)
print("ANÁLISIS DE INVERSIONES VS RETIROS POR USUARIO")
print("=" * 80)
print()

# Crear workbook de Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Análisis Inversiones"

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezados
headers = [
    "ID Usuario",
    "Username",
    "Email",
    "Total Invertido",
    "Total Retirado",
    "P2P Enviado",
    "Balance",
    "Estado",
    "# Depósitos",
    "# Retiros",
    "# Transferencias P2P",
    "UserInvestment.withdrawn",
    "Suma withdrawn_from_deposit",
    "Inconsistencia withdrawn",
    "Inconsistencia deposits"
]

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 25
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 20
ws.column_dimensions['L'].width = 22
ws.column_dimensions['M'].width = 25
ws.column_dimensions['N'].width = 25
ws.column_dimensions['O'].width = 25

# Obtener todos los usuarios con inversiones
user_investments = UserInvestment.objects.filter().select_related('user')

total_users = 0
users_with_excess_withdrawal = 0
users_with_balance = 0
current_row = 2

for user_investment in user_investments:
    user = user_investment.user
    total_users += 1
    
    print(f"Procesando: {user.username}...")
    
    # 1. Calcular total invertido (suma de todos los deposits aprobados)
    approved_deposits = InvestmentPlanTransaction.objects.filter(
        current_investment=user_investment,
        status=TransactionStatus.APPROVED
    )
    
    total_invested = Decimal('0')
    deposit_count = 0
    
    for deposit in approved_deposits:
        total_invested += deposit.amount
        deposit_count += 1
    
    # 2. Calcular total retirado (suma de todos los retiros de inversión aprobados)
    investment_withdrawals = Withdrawal.objects.filter(
        user=user,
        type=WithdrawalType.INVESTMENT,
        status__in=[TransactionStatus.APPROVED, TransactionStatus.PENDING]
    )
    
    total_withdrawn = Decimal('0')
    withdrawal_count = 0
    
    for withdrawal in investment_withdrawals:
        total_withdrawn += withdrawal.amount
        withdrawal_count += 1
    
    # 3. Calcular total enviado en transferencias P2P de inversión
    p2p_transfers = UserTransfer.objects.filter(
        sender=user,
        transfer_type=UserTransfer.TransferType.INVESTMENT
    )
    
    total_p2p_sent = Decimal('0')
    p2p_count = 0
    
    for transfer in p2p_transfers:
        total_p2p_sent += transfer.amount
        p2p_count += 1
    
    # 4. Calcular balance (Total Invertido - Total Retirado - P2P Enviado)
    balance = total_invested - total_withdrawn - total_p2p_sent
    
    # 5. Determinar estado
    if balance < 0:
        estado = "⚠️ HA RETIRADO DE MÁS"
        users_with_excess_withdrawal += 1
        row_fill = error_fill
    else:
        estado = "✅ ESTÁ BIEN"
        users_with_balance += 1
        row_fill = ok_fill
    
    # 6. Calcular withdrawn_from_deposit total
    total_withdrawn_from_deposits = Decimal('0')
    for deposit in approved_deposits:
        total_withdrawn_from_deposits += deposit.withdrawn_from_deposit
    
    # 7. Verificar inconsistencias
    inconsistencia_withdrawn = "Sí" if user_investment.withdrawn != total_withdrawn else "No"
    inconsistencia_deposits = "Sí" if total_withdrawn != total_withdrawn_from_deposits else "No"
    
    # Aplicar color de advertencia si hay inconsistencias
    if inconsistencia_withdrawn == "Sí" or inconsistencia_deposits == "Sí":
        if balance >= 0:  # Solo si no es error grave
            row_fill = warning_fill
    
    # 8. Escribir en Excel
    data = [
        user.id,
        user.username,
        user.email,
        float(total_invested),
        float(total_withdrawn),
        float(total_p2p_sent),
        float(balance),
        estado,
        deposit_count,
        withdrawal_count,
        p2p_count,
        float(user_investment.withdrawn),
        float(total_withdrawn_from_deposits),
        inconsistencia_withdrawn,
        inconsistencia_deposits
    ]
    
    for col_num, value in enumerate(data, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = value
        cell.border = border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formato de números
        if col_num in [4, 5, 6, 7, 12, 13]:  # Columnas de dinero (Total Invertido, Total Retirado, P2P Enviado, Balance, withdrawn, suma withdrawn_from_deposit)
            cell.number_format = '$#,##0.00'
    
    current_row += 1

# Resumen final
print(f"\n\n{'='*80}")
print(f"RESUMEN GENERAL")
print(f"{'='*80}")
print(f"Total de usuarios analizados: {total_users}")
print(f"Usuarios con retiros excesivos (balance negativo): {users_with_excess_withdrawal}")
print(f"Usuarios con balance correcto (>= 0): {users_with_balance}")
print(f"{'='*80}\n")

# Guardar archivo Excel
filename = f"analisis_inversiones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(filename)

print(f"\n✅ Archivo Excel generado: {filename}")
print(f"   Ubicación: {filename}")
print(f"\n📊 Estadísticas:")
print(f"   - Total usuarios: {total_users}")
print(f"   - ✅ Están bien: {users_with_balance}")
print(f"   - ⚠️  Han retirado de más: {users_with_excess_withdrawal}")
print(f"\n{'='*80}\n")
